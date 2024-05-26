import csv
import random 
from django.core.cache import cache
from django.shortcuts import render,redirect,HttpResponse
# from django.contrib.auth.models import User
from django.contrib.auth import logout,authenticate,login
# from accounts.models import Attendance
from django.contrib import messages
from django.core.mail import send_mail
from django.conf import settings
import openpyxl.utils
import openpyxl.utils.exceptions
# from .models import User_Data
from accounts.models import CustomUser
import openpyxl       # for storind datas using excel format
from attendances.models import AttendanceData
from academics.models import *   # import academics model



# Create your views here.

# =======user authenitcation======

def login_view(request):
    if request.method =='POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user_type = request.POST.get('user_type')
        
        try:
            is_a_user = CustomUser.objects.filter(username = username, user_type = user_type).exists()
            user = authenticate(request, username=username, password=password)
            print(user)
            if user is not None and is_a_user:
                login(request, user)
                if user_type == 'student':
                    return redirect('academics:student_dashboard')
                elif user_type == 'teacher':
                    return redirect('academics:teacher_dashboard')    
            else:
                error_message = 'Invalid username or password'
                return render(request, 'accounts/login.html', {'error_message': error_message})
        except CustomUser.DoesNotExist:
            error_message = 'Invalid username or user type'
            return render(request, 'accounts/login.html', {'error_message': error_message})
    else:
        return render(request, 'accounts/login.html')


# =========forgot-password=======

def forgot_password(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        try:
            user = CustomUser.objects.get(username = username, email = email)
            request.session['user_email'] = user.email
            request.session['username'] = user.username
            
            otp = str(random.randint(111111, 999999))
            print(otp)
            cache_key = f'otp_{user.email}'
            cache.set(cache_key, otp, timeout=600)  # OTP expires in 10 minutes
            subject = "OTP for password reset"
            message = f"Your OTP is: {otp}\n\nPlease note that this OTP will expire in 10 minutes."
            from_email = settings.EMAIL_HOST_USER
            recipient_email = [user.email]
            send_mail(subject,message,from_email,recipient_email)
            # message = "Otp Sent Successfully!"
            return redirect('accounts:otp-verification')
        except CustomUser.DoesNotExist:
            error_message = "Wrong username or email!"
            return render(request, 'accounts/forgot_password.html', {'message': error_message})   
    return render(request, 'accounts/forgot_password.html')

# ======otp verifying=========

def otp_verifying(request):
    print("hel")
    if request.method == 'POST':
        entered_otp = request.POST.get('otp')
        email = request.session.get('user_email')
        username = request.session.get('username')
        print(username)
        print(email)
        cache_key =  f'otp_{email}'
        stored_otp = cache.get(cache_key)
        if stored_otp is not None:
            if entered_otp == stored_otp:
                
                return render(request, 'accounts/otp_and_password_reset.html', {'success': True})
            else:
                error_message = "Incorrect otp. Please try again!"
                return render(request, 'accounts/otp_and_password_reset.html', {'success': False, 'error_message':error_message})
        else:
            # error_message = "Otp expired!"
            return redirect('accounts:forgot-password')
    else:
        return render(request, 'accounts/otp_and_password_reset.html')
    

# =======password reset======

def change_password(request):
    print("hello")
    if request.method == 'POST':
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')
        username = request.session.get('username')
        print(username)
        try:
            user = CustomUser.objects.get(username = username)
            if password == confirm_password:
                user.set_password(confirm_password)
                user.save()
                print('Password changed successfuly!')
                return redirect('accounts:login')
            else:
                error_message = "Password do not match. Please try again"
                return render(request, 'accounts/otp_and_password_reset.html', {'error_message':error_message ,'success':True} )
        
        except CustomUser.DoesNotExist:
            error_message = "User not found."
            return render(request, 'accounts/otp_and_password_reset.html', {'error_message': error_message, 'success': True})
    else:
        return redirect('accounts:forgot-password')


# ======upload data======
def upload(request):
    if request.user.is_authenticated and request.user.is_superuser:
        if request.method == 'POST':
            if 'users' in request.FILES:
                uploaded_file = request.FILES['users']
                try:
                    # Load the Excel workbook
                    workbook = openpyxl.load_workbook(uploaded_file)
                    sheet = workbook.active

                    # Iterate over each row in the sheet (starting from the second row)
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        username = row[0]
                        password = row[1]
                        name = row[2] if len(row) > 2 else '0'
                        email = row[3] if len(row) > 3 else 'null'
                        gender = row[4] if len(row) > 4 else '0'
                        user_type = row[5] if len(row) > 5 else '0'
                        course_id = row[6] if len(row) > 6 else '0'
                        year_id = row[7] if len(row) > 7 else '0'
                        sem = row[8] if len(row) > 8 else '0'
                        language_pref = row[9] if len(row) > 9 else '0'
                        course_type = row[10] if len(row) > 10 else '0'
                        section_id = row [11] if len(row) > 11 else '0'
                        admission_year = row[12] if len(row) > 12 else 0
                        lab_batch_type = row[13] if len(row) > 13 else '0'

                        # # Check if a user with the same username already exists
                        # if CustomUser.objects.filter(username=username).exists():
                        #     continue  # Skip the row and move on to the next one

                        # Perform data validation
                        if not username or not password:
                            raise ValueError("Username and password are required.")
                
                        if user_type not in ['admin', 'student', 'teacher']:
                            raise ValueError("Invalid user type.")
                
                        if gender not in ['m', 'f']:
                            raise ValueError("Invalid gender.")
                
                        if sem not in ['0','1','2','3','4','5','6', 'even', 'odd', 'passed out']:
                            raise ValueError("Invalid semester.")
                
                        if language_pref not in ['0','hindi', 'kannada', 'malayalam']:
                            raise ValueError("Invalid language preference.")
                        
                        if course_type not in ['normal','ai/ml','logistics','aviation','0']:
                            raise ValueError("Invalid batch type.")
                        
                        if lab_batch_type not in ['0', 'normal', 'addon']:
                            raise ValueError("Invalid lab batch.")
                        
                        if section_id not in ['0','A','B']:
                            raise ValueError("Invalid Section!")
                        
                        # Check if the user already exists in the database
                        is_user= CustomUser.objects.filter(username=username).exists()
                        

                        if is_user:
                        # User already exists, update the fields
                            user = CustomUser.objects.get(username=username)
                            user.name = name
                            user.email = email
                            user.gender = gender
                            user.user_type = user_type
                            user.course_id = course_id
                            user.year_id = year_id
                            user.sem = sem
                            user.language_pref = language_pref
                            user.course_type = course_type
                            user.section_id = section_id
                            user.AdmissionYear = admission_year
                            user.lab_batch_type = lab_batch_type
                            user.save()

                        else:
                            # Create a new user with the provided data
                            user = CustomUser(
                                username=username,
                                name=name,
                                email=email,
                                gender=gender,
                                user_type=user_type,
                                course_id=course_id,
                                year_id=year_id,
                                sem=sem,
                                language_pref=language_pref,
                                course_type=course_type,
                                section_id=section_id,
                                AdmissionYear=admission_year,
                                lab_batch_type=lab_batch_type
                            )
                            user.set_password(password)  # Use set_password() to set the password
                            user.save()
                    return redirect('accounts:upload')

                except (ValueError, openpyxl.utils.exceptions.InvalidFileException) as e:
                    error_message = str(e)
                    return render(request, 'accounts/upload_files.html', {'error_message': error_message})
            elif 'course' in request.FILES:
                uploaded_file = request.FILES['course']
                try:
                    # Load the Excel workbook
                    workbook = openpyxl.load_workbook(uploaded_file)
                    sheet = workbook.active

                    # Iterate over each row in the sheet (starting from the second row)
                    for row in sheet.iter_rows(min_row = 2, values_only = True):
                        course_id = row[0]
                        course_name = row[1]

                        # Check if the course already exists
                        if Course.objects.filter(course_id = course_id).exists():           # if the course_id is already present
                            continue  # Skip the row if the course already exists           # in the excel sheet then it will be skipped

                        course = Course(course_id = course_id, course_name = course_name)
                        course.save()
                    return redirect('accounts:upload')
                
                except (ValueError,openpyxl.utils.exceptions.InvalidFileException) as e:
                    error_message = str(e)
                    return render(request, 'accounts/upload_files.html', {'error_message':error_message})
            elif 'year' in request.FILES:
                uploaded_file = request.FILES['year']
                try:
                    # Load the Excel workbook
                    workbook = openpyxl.load_workbook(uploaded_file)
                    sheet = workbook.active

                    # Iterate over each row in the sheet (starting from the second row)
                    for row in sheet.iter_rows(min_row = 2, values_only = True):
                        year_id = row[0]
                        year_name = row[1]
                        course_id = row[2]

                        if Year.objects.filter(year_id = year_id).exists():  # if year_id is already there in the database then
                            continue                                         # the entire row from the excel sheet will be skipped

                        course = Course.objects.get(course_id=course_id)
                        year = Year(year_id=year_id, year_name=year_name, course_id=course)
                        year.save()
                    return redirect('accounts:upload')
                except (ValueError, openpyxl.utils.exceptions.InvalidFileException, Course.DoesNotExist) as e:
                    error_message = str(e)
                    return render(request, 'accounts/upload_files.html', {'error_message': error_message})
            # elif 'section' in request.FILES:
            #     uploaded_file = request.FILES['section']
            #     try:
            #         # Load the Excel workbook
            #         workbook = openpyxl.load_workbook(uploaded_file)
            #         sheet = workbook.active

            #         # Iterate over each row in the sheet (starting from the second row)
            #         for row in sheet.iter_rows(min_row = 2, values_only = True):
            #             section_id = row[0]
            #             section_name = row[1]
            #             year_id = row[2]
            #             course_id = row[3]

            #             if Section.objects.filter(section_id = section_id).exists():
            #                 continue
            #             year = Year.objects.get(year_id=year_id)
            #             course = Course.objects.get(course_id=course_id)
            #             section = Section(section_id=section_id, section_name=section_name, year_id=year, course_id=course)
            #             section.save()
            #     except (ValueError, openpyxl.utils.exceptions.InvalidFileException, Year.DoesNotExist, Course.DoesNotExist) as e:
            #         error_message = str(e)
            #         return render(request, 'accounts/upload_files.html', {'error_message': error_message})
            elif 'subject' in request.FILES:    
                uploaded_file = request.FILES['subject']
                try:
                    # Load the Excel workbook
                    workbook = openpyxl.load_workbook(uploaded_file)
                    sheet = workbook.active

                    # Iterate over each row in the sheet (starting from the second row)
                    for row in sheet.iter_rows(min_row = 2, values_only = True):
                        subject_code = row[0]
                        subject_name = row[1]
                        subject_type = row[2]
                        sem = row[3]
                        year_id = row[4]
                        teacher_id = row[5]
                        section = row[6] if len(row) > 6 else '0'  # Get section or default to '0'
                        lab_batch_type = row[7] if len(row) > 7 else '0'  # Get lab_batch_type if no batch found then store 0

                        if Subject.objects.filter(subject_code=subject_code,subject_name=subject_name, section = section, teacher_id = teacher_id, lab_batch_type=lab_batch_type).exists():
                            continue #skip

                        try:
                            print(year_id)
                            course_name= Year.objects.get(year_id=year_id)
                            print("####\n",year_id)
                            teacher = CustomUser.objects.get(username=teacher_id)
                        except (Course.DoesNotExist, CustomUser.DoesNotExist):
                            error_message = f"Invalid course ID or teacher ID: {year_id}, {teacher_id}"
                            return render(request, 'accounts/upload_files.html', {'error_message': error_message})

                        subject = Subject(
                            subject_code=subject_code,
                            subject_name=subject_name,
                            year_id=course_name,
                            teacher_id=teacher,
                            subject_type=subject_type,
                            sem=sem,
                            section = section,
                            lab_batch_type=lab_batch_type
                            )
                        subject.save()

                except (ValueError, openpyxl.utils.exceptions.InvalidFileException, Course.DoesNotExist, CustomUser.DoesNotExist) as e:
                    error_message = str(e)
                    return render(request, 'accounts/upload_files.html', {'error_message': error_message})
            # elif 'teachersubject' in request.FILES:  
            #     uploaded_file = request.FILES['section']
            #     try:
            #         # Load the Excel workbook
            #         workbook = openpyxl.load_workbook(uploaded_file)
            #         sheet = workbook.active

            #         # Iterate over each row in the sheet (starting from the second row)
            #         for row in sheet.iter_rows(min_row = 2, values_only = True):
            #             teacher_id = row[0]
            #             subject_code = row[1]
            #             course_id = row[2]
            #             sem = row[3]    

            #             user = CustomUser.objects.get(username = teacher_id)
            #             subject = Subject.objects.get(subject_code=subject_code)
            #             course = Course.objects.get(course_id=course_id)
            #             teacher_subject = TeacherSubject(teacher_id=user, subject_code=subject, course_id=course, sem=sem)
            #             teacher_subject.save()

            #     except (ValueError, openpyxl.utils.exceptions.InvalidFileException, Subject.DoesNotExist, Course.DoesNotExist) as e:
            #         error_message = str(e)
            #         return render(request, 'academics/upload_files.html', {'error_message': error_message})

            return render(request, 'accounts/upload_files.html')
        
        return render(request, 'accounts/upload_files.html')

    elif not request.user.is_anonymous and CustomUser.objects.filter(username = request.user.username, user_type = 'student').exists():
         return redirect('academics:student_dashboard')
    elif not request.user.is_anonymous and CustomUser.objects.filter(username = request.user.username, user_type = 'teacher').exists():
        return redirect('academics:teacher_dashboard')
    elif request.user.is_anonymous:
        return redirect('accounts:login')

        

        
























# ==================home===================

def home(request):
    if request.user.is_anonymous:
        return redirect('accounts/login.html')
    return render(request,'home.html')


# ==============check login credentials==============

def loginUser(request):
    if request.method =='POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        print(username,password)
        #check user has entered correct credentials
        user = authenticate(username=username, password=password)
        
        # checks if user added credentials or not
        if user is not None:
            login(request, user)
            messages.success(request, "Login Successful!")
            return redirect("/home")
        else:
            return render(request, "login.html")
            

    return render(request,'login.html')

def logoutUser(request):
    logout(request)
    return redirect('/login')

# ================attendance csvfile upload======================= 
#============== dont need it for right now============


def uploadCsvFile(request):
    if request.method == 'POST' and request.FILES['csv_file']:
        csv_file = request.FILES['csv_file']
        decoded_file = csv_file.read().decode('utf-8').splitlines()
        reader = csv.DictReader(decoded_file)
        for row in reader:
                attendance = AttendanceData(
                    registration_no=row['Registration No'],
                    section_id=row['Section ID'],
                    subject_id=row['Subject ID'],
                    date=row['Date'],
                    period=int(row['Period']),  # Convert to integer
                    status=row['Status'],
                    course_id=row['Course ID']
                )
                attendance.save()
        return HttpResponse("Succesfull!")
    return render(request, 'upload_csv.html')

# ==============check username and email matches with the database=================

# def check_Username_Password_send_Otp(request):
#     if request.method == 'POST':
#         username = request.POST.get('username')
#         email = request.POST.get( 'email' )
#         try:
#             # checking if user is found or not 
#             user = User.objects.get(username=username , email =email)
#             print(user)
#              # Generate OTP
#             otp = str(random.randint(111111, 999999))
#             print(otp)
#             cache_key = f'otp_{email}'
#             print(cache_key)
#             cache.set(cache_key, otp, timeout=600)  # OTP expires in 10 minutes
#             subject = "OTP for password reset"
#             message = f"Your OTP is: {otp}\n\nPlease note that this OTP will expire in 10 minutes."
#             from_email = settings.EMAIL_HOST_USER
#             recipient_email = [email]
#             send_mail(subject,message,from_email,recipient_email)
            

#             return redirect("verifying otp", username=username)
        
#         except User.DoesNotExist: 
#              error_message = 'Invalid username or email.'
#              return render(request, 'forgot_password.html', {'error_message': error_message})

#     return render(request, 'forgot_password.html')

# ==============check if the otp matches with the genrated otp ===================

# def verify_otp(request,username):
#     if request.method == 'POST':
#         try:
#             user = User.objects.get(username = username)
#             user_entered_otp = request.POST.get( 'otp' )
#             cache_key =  f'otp_{user.email}'
#             stored_otp = cache.get(cache_key)
#             print(user_entered_otp)

#             if stored_otp is not None:
#                 if user_entered_otp == stored_otp:
#                     cache.delete(cache_key)
#                     return redirect("password_reset.html",username= username)
#                 else:
#                     error_message = 'Invalid OTP. Please try again.'
#                     return render(request ,'verifying otp',{'error_message':error_message , 'username':username} )
#             else:
#                 error_message = 'OTP has expired. Please request a new OTP.'
#                 return render(request ,'check data and send otp',{'error_message':error_message} )

#         except User.DoesNotExist:
#             error_message = 'User does not exist.'
#         return render(request, 'verifying otp',{'error_message': error_message ,'username': username})

#     return render(request, 'otp_verification.html', {'username': username})

# ============reseting the password =============

# def reset_password(request,username):
#     if request.method == 'POST':
#         changed_password = request.POST.get('password')
#         confirm_changed_password = request.POST.get('confirmPassword')
#         try:
#             if changed_password == confirm_changed_password:
#                 user =User.objects.get(username = username)
#                 user.set_password(changed_password)
#                 user.save()
#                 return redirect('login_page')
#             else:
#                 error_message = "Password do not Match!"
#                 return render(request, 'password_reset',{'error_message':error_message , 'username':username })
#         except User.DoesNotExist:
#             error_message = "User does not exist!"
#             return render(request, 'check data and send otp' , {'error_message':error_message})
#         except Exception as e:
#             error_message = "An error occurred during password reset. Please try again."
#             return render(request , 'password_reset', {'error_message':error_message , 'username':username })
              
#     return render (request,'password_reset')